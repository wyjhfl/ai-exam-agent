import io
import logging
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, Integer
from db.database import get_session
from db.models import User, WrongQuestion, QuizQuestion, QuizRecord, StudySession, StudyPlan
from models.schemas import UserInfo

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/{user_id}/wrong-questions")
async def export_wrong_questions(user_id: int, session: AsyncSession = Depends(get_session)):
    result = await session.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    result = await session.execute(
        select(WrongQuestion, QuizQuestion)
        .join(QuizQuestion, WrongQuestion.question_id == QuizQuestion.id)
        .where(WrongQuestion.user_id == user_id)
    )
    rows = result.all()

    items = []
    for wq, q in rows:
        items.append({
            "subject": q.subject,
            "question": q.question_text,
            "options": q.options,
            "your_answer": "",
            "correct_answer": q.answer,
            "explanation": q.explanation,
            "review_count": wq.review_count,
            "mastered": wq.mastered,
        })

    result = await session.execute(
        select(QuizRecord, QuizQuestion)
        .join(QuizQuestion, QuizRecord.question_id == QuizQuestion.id)
        .where(QuizRecord.user_id == user_id, QuizRecord.is_correct == False)
    )
    records = result.all()
    for rec, q in records:
        already = any(it["question"] == q.question_text for it in items)
        if not already:
            items.append({
                "subject": q.subject,
                "question": q.question_text,
                "options": q.options,
                "your_answer": rec.selected_answer or "",
                "correct_answer": q.answer,
                "explanation": q.explanation,
                "review_count": 0,
                "mastered": False,
            })

    return {"user_id": user_id, "username": user.username, "wrong_questions": items}


@router.get("/{user_id}/study-summary")
async def export_study_summary(user_id: int, session: AsyncSession = Depends(get_session)):
    result = await session.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    total_quiz = await session.execute(
        select(func.count()).where(QuizRecord.user_id == user_id)
    )
    total_quiz_count = total_quiz.scalar() or 0

    correct_quiz = await session.execute(
        select(func.count()).where(QuizRecord.user_id == user_id, QuizRecord.is_correct == True)
    )
    correct_count = correct_quiz.scalar() or 0

    accuracy = round(correct_count / total_quiz_count * 100, 1) if total_quiz_count > 0 else 0

    focus_result = await session.execute(
        select(func.coalesce(func.sum(StudySession.duration), 0)).where(
            StudySession.user_id == user_id, StudySession.session_type == "focus"
        )
    )
    total_focus_minutes = focus_result.scalar() or 0

    subject_result = await session.execute(
        select(QuizQuestion.subject, func.count(), func.sum(
            func.cast(QuizRecord.is_correct, Integer)
        ))
        .join(QuizQuestion, QuizRecord.question_id == QuizQuestion.id)
        .where(QuizRecord.user_id == user_id)
        .group_by(QuizQuestion.subject)
    )
    subject_stats = []
    for subject, total, correct in subject_result.all():
        subject_stats.append({
            "subject": subject,
            "total": total,
            "correct": correct or 0,
            "accuracy": round((correct or 0) / total * 100, 1) if total > 0 else 0,
        })

    plan_result = await session.execute(
        select(StudyPlan).where(StudyPlan.user_id == user_id, StudyPlan.is_active == True)
    )
    active_plan = plan_result.scalar_one_or_none()

    return {
        "user_id": user_id,
        "username": user.username,
        "total_quiz_count": total_quiz_count,
        "correct_count": correct_count,
        "accuracy": accuracy,
        "total_focus_minutes": total_focus_minutes,
        "subject_stats": subject_stats,
        "has_active_plan": active_plan is not None,
    }


@router.get("/{user_id}/wrong-questions/excel")
async def export_wrong_questions_excel(user_id: int, session: AsyncSession = Depends(get_session)):
    from openpyxl import Workbook

    data = await export_wrong_questions(user_id, session)

    wb = Workbook()
    ws = wb.active
    ws.title = "错题本"
    headers = ["科目", "题目", "你的答案", "正确答案", "解析", "复习次数", "掌握状态"]
    ws.append(headers)

    for item in data["wrong_questions"]:
        ws.append([
            item["subject"],
            item["question"],
            item["your_answer"],
            item["correct_answer"],
            item["explanation"] or "",
            item["review_count"],
            "已掌握" if item["mastered"] else "未掌握",
        ])

    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=wrong_questions_{user_id}.xlsx"},
    )


@router.get("/{user_id}/study-summary/excel")
async def export_study_summary_excel(user_id: int, session: AsyncSession = Depends(get_session)):
    from openpyxl import Workbook

    data = await export_study_summary(user_id, session)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "学习总览"
    ws1.append(["指标", "数值"])
    ws1.append(["用户名", data["username"]])
    ws1.append(["总刷题数", data["total_quiz_count"]])
    ws1.append(["正确数", data["correct_count"]])
    ws1.append(["正确率", f'{data["accuracy"]}%'])
    ws1.append(["总专注时长(分钟)", data["total_focus_minutes"]])
    ws1.append(["是否有活跃计划", "是" if data["has_active_plan"] else "否"])

    for col in ws1.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws1.column_dimensions[col[0].column_letter].width = max_length + 2

    ws2 = wb.create_sheet(title="各科统计")
    ws2.append(["科目", "总题数", "正确数", "正确率"])
    for stat in data["subject_stats"]:
        ws2.append([stat["subject"], stat["total"], stat["correct"], f'{stat["accuracy"]}%'])

    for col in ws2.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws2.column_dimensions[col[0].column_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=study_summary_{user_id}.xlsx"},
    )


@router.get("/{user_id}/wrong-questions/pdf")
async def export_wrong_questions_pdf(user_id: int, session: AsyncSession = Depends(get_session)):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    FONT_NAME = "Helvetica"
    font_paths = [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simsun.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    ]
    for fp in font_paths:
        try:
            pdfmetrics.registerFont(TTFont("ChineseFont", fp))
            FONT_NAME = "ChineseFont"
            break
        except Exception:
            pass

    data = await export_wrong_questions(user_id, session)
    username = data.get("username", "用户")
    items = data.get("wrong_questions", [])

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    margin = 2 * cm
    usable_w = w - 2 * margin
    y = h - margin

    def draw_footer(page_num):
        c.setFont(FONT_NAME, 8)
        c.drawString(margin, 1 * cm, f"第 {page_num} 页")
        c.drawRightString(w - margin, 1 * cm, f"导出日期: {datetime.now().strftime('%Y-%m-%d')}")

    def new_page(page_num):
        if page_num > 1:
            draw_footer(page_num - 1)
            c.showPage()
        return h - margin

    page_num = 1
    y = new_page(page_num)

    c.setFont(FONT_NAME, 18)
    c.drawString(margin, y, f"错题本 - {username}")
    y -= 1.5 * cm

    c.setFont(FONT_NAME, 10)
    c.drawString(margin, y, f"共 {len(items)} 道错题")
    y -= 1.5 * cm

    items_on_page = 0

    for idx, item in enumerate(items, 1):
        if items_on_page >= 3 or y < 6 * cm:
            page_num += 1
            y = new_page(page_num)
            items_on_page = 0

        c.setFont(FONT_NAME, 12)
        c.drawString(margin, y, f"{idx}. [{item.get('subject', '')}]")
        y -= 0.8 * cm

        c.setFont(FONT_NAME, 10)
        question_text = item.get("question", "")
        max_chars = int(usable_w / 6)
        for start in range(0, len(question_text), max_chars):
            c.drawString(margin, y, question_text[start:start + max_chars])
            y -= 0.6 * cm

        options = item.get("options", [])
        if options:
            for opt in options[:4]:
                opt_text = str(opt)
                c.drawString(margin + 0.5 * cm, y, opt_text[:max_chars])
                y -= 0.5 * cm

        c.setFont(FONT_NAME, 10)
        correct = item.get("correct_answer", "")
        c.drawString(margin, y, f"正确答案: {correct}")
        y -= 0.6 * cm

        explanation = item.get("explanation", "")
        if explanation:
            c.setFont(FONT_NAME, 9)
            for start in range(0, len(explanation), max_chars):
                c.drawString(margin + 0.5 * cm, y, f"解析: {explanation[start:start + max_chars]}")
                y -= 0.5 * cm

        y -= 0.8 * cm
        items_on_page += 1

    draw_footer(page_num)
    c.save()
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=wrong_questions_{user_id}.pdf"},
    )
